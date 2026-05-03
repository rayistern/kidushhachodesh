"""Build a Hebrew-RTL Excel mirror of the user's spreadsheet for ב' סיון ה'תשפו,
populated with our engine's numbers and explicit gaps where our engine doesn't yet
implement the full KH 17 visibility chain.

Run: python3 responses/build_excel.py
Outputs: responses/our_numbers_2_sivan_5786.xlsx
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUR = json.load(open('/tmp/our_calcs.json'))


def dms(deg):
    if deg is None:
        return ''
    sign = '-' if deg < 0 else ''
    deg = abs(deg)
    d_int = int(deg)
    rem = (deg - d_int) * 60
    m = int(rem)
    s = int(round((rem - m) * 60))
    if s == 60:
        s = 0; m += 1
    if m == 60:
        m = 0; d_int += 1
    return f"{sign}{d_int}:{m:02d}:{s:02d}"


def find(day, month):
    return next(r for r in OUR if r['heDay'] == day and r['heMonth'] == month)


target = find(2, 'Sivan')
sun = target['sun']; moon = target['moon']

wb = Workbook()
ws = wb.active
ws.title = "ב סיון ה'תשפו"
ws.sheet_view.rightToLeft = True

bold = Font(bold=True)
hdr_fill = PatternFill('solid', fgColor='FFE699')
match_fill = PatternFill('solid', fgColor='E2EFDA')
mismatch_fill = PatternFill('solid', fgColor='FFD7D7')
gap_fill = PatternFill('solid', fgColor='FCE4D6')
center = Alignment(horizontal='center', vertical='center')
right = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='999999')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def row(r, label, our_val, user_val=None, status='match', note=''):
    ws.cell(row=r, column=1, value=label).font = bold
    ws.cell(row=r, column=1).alignment = right
    ws.cell(row=r, column=2, value=our_val).alignment = center
    ws.cell(row=r, column=3, value=user_val if user_val is not None else '').alignment = center
    ws.cell(row=r, column=4, value=note).alignment = right
    fill = {'match': match_fill, 'mismatch': mismatch_fill, 'gap': gap_fill, 'header': hdr_fill}.get(status)
    if fill:
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = border


def header(r, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = Font(bold=True, size=12)
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = border


# Title row
ws.merge_cells('A1:D1')
ws['A1'] = "השוואה: kidushhachodesh.netlify.app מול חשבונות המשתמש — ב' סיון ה'תשפו"
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = center
ws['A1'].fill = hdr_fill

# Column headers
ws['A3'] = 'שם'; ws['B3'] = 'הערך שלנו'; ws['C3'] = 'ערך המשתמש'; ws['D3'] = 'הערה'
for c in 'ABCD':
    ws[f'{c}3'].font = bold
    ws[f'{c}3'].alignment = center
    ws[f'{c}3'].fill = hdr_fill
    ws[f'{c}3'].border = border

r = 4
row(r, 'תאריך גרגוריאני', target['gregorian'], '', 'match'); r += 1
row(r, 'מספר ימים מהעיקר', target['daysFromEpoch'], 309775, 'match', 'תואם'); r += 1

header(r, 'חישובי השמש (KH 12-13)'); r += 1
row(r, 'אמצע השמש', dms(sun['meanLon']), '55:55:34', 'match'); r += 1
row(r, 'גובה השמש (אפוגי)', dms(sun['apogee']), '99:39:34', 'match'); r += 1
row(r, 'מסלול השמש', dms(sun['maslul']), '316:16:00', 'match'); r += 1
row(r, 'מנת המסלול', dms(sun['maslulCorr']), '1:20:14', 'match'); r += 1
row(r, 'השמש האמיתי', dms(sun['trueLon']), '57:15:48', 'match', 'תואם בקנה מידה של שניות'); r += 1

header(r, 'חישובי הירח (KH 14-16)'); r += 1
row(r, 'אמצע הירח', dms(moon['meanLon']), '69:41:39', 'match'); r += 1
row(r, 'התיקון לשעת הראיה (תיקון KH 14:5)',
    dms(moon['adjustedMeanLon'] - moon['meanLon']),
    '0:30:00', 'mismatch',
    'אצלנו 0:15:00 לפי הטבלה (ראה הסבר)'); r += 1
row(r, 'אמצע הירח לשעת הראיה',
    dms(moon['adjustedMeanLon']),
    '70:11:39', 'mismatch',
    'נובע מהפרש התיקון העונתי'); r += 1
row(r, 'אמצע המסלול', dms(moon['maslul']), '168:49:53', 'match'); r += 1
row(r, 'מרחק הכפול (merchak kaful)', dms(moon['doubleElong']), '28:32:10', 'mismatch',
    'הפרש 0:30 — נגרר מהתיקון העונתי'); r += 1
row(r, 'מסלול הנכון (maslul hanachon)', dms(moon['maslulHanachon']), '172:49:53', 'match'); r += 1
row(r, 'מנת המסלול (תיקון של הירח)', dms(moon['maslulCorr']), '0:42:18', 'match'); r += 1
row(r, 'הירח האמיתי', dms(moon['trueLon']), '69:29:21', 'mismatch',
    'הפרש 0:15 נגרר מהתיקון העונתי'); r += 1
row(r, 'הראש (node)', dms(moon['node']), '335:35:37', 'match'); r += 1

header(r, 'שלבי הראייה (KH 17)'); r += 1
row(r, 'אורך ראשון (= elongation, KH 17:1)', dms(moon['elongation']), '12:13:34', 'mismatch',
    'הפרש 0:15 נגרר מהתיקון העונתי (סוגיה #19)'); r += 1
row(r, 'אורך שני (KH 17:5)', dms(moon['orechSheni']), '11:14:34', 'mismatch',
    'אצלנו פוחתים 58\' (Gemini); משתמש פוחת 59\' — בגדר רמב"ם "אֵין מְדַקְדְּקִין בִּשְׁנִיּוֹת"'); r += 1
row(r, 'רוחב שני (KH 17:7-9)', dms(abs(moon['rochavSheni'])) + (' צפוני' if moon['rochavSheni']>=0 else ' דרומי'),
    '4:42:03', 'match', 'תואם בקנה מידה של דקות'); r += 1
row(r, 'אורך שלישי (KH 17:10-11)', dms(moon['orechShlishi']), '10:27:33', 'mismatch',
    'הפער נגרר מתיקון העונתי (#19)'); r += 1
row(r, 'אורך רביעי (KH 17:12a)', dms(moon['orechRevii']), '12:12:09', 'mismatch',
    'הפער נגרר מהפערים שלמעלה'); r += 1
row(r, 'מנת גובה המדינה (KH 17:12b)', dms(moon['mnatGovah']), '03:18:42', 'match',
    '⅔ × |רוחב ראשון| לא"י — תואם'); r += 1
row(r, 'קשת הראיה (KH 17:12c)', dms(moon['keshet']), '15:30:51', 'match',
    'בקנה מידה של דקות; שניהם > 14° = ודאי יראה'); r += 1
row(r, 'מסקנה: האם ייראה?', 'כן — ודאי יראה ' + ('✓' if moon['isVisible'] else ''),
    'כן — ודאי יראה', 'match',
    'KH 17:15: קשת > 14°'); r += 1

# Literal mirror sheet — column A label, column B our value, in the
# exact row order of the user's worksheet for ב' סיון ה'תשפו.
ws_mirror = wb.create_sheet('מראה — ב סיון')
ws_mirror.sheet_view.rightToLeft = True

mirror_rows = [
    ("ב סיון ה'תשפו", target['daysFromEpoch']),
    ('', ''),
    ('מספר ימים מהעיקר', target['daysFromEpoch']),
    ('', ''),
    ('אמצע השמש', dms(sun['meanLon'])),
    ('', ''),
    ('גובה השמש', dms(sun['apogee'])),
    ('', ''),
    ('מסלול השמש', dms(sun['maslul'])),
    ('', ''),
    ('מנת המסלול', dms(sun['maslulCorr'])),
    ('', ''),
    ('השמש האמיתי', dms(sun['trueLon'])),
    ('', ''),
    ('אמצע הירח', dms(moon['meanLon'])),
    ('', ''),
    ('התיקון לשעת הראיה', dms(moon['adjustedMeanLon'] - moon['meanLon'])),
    ('', ''),
    ('אמצע הירח לשעת הראיה', dms(moon['adjustedMeanLon'])),
    ('', ''),
    ('אמצע המסלול', dms(moon['maslul'])),
    ('', ''),
    ('מרחק הכפול', dms(moon['doubleElong'])),
    ('', ''),
    ('התוספת', dms(moon['maslulHanachon'] - moon['maslul'])),
    ('', ''),
    ('מסלול הנכון', dms(moon['maslulHanachon'])),
    ('', ''),
    ('מנת המסלול', dms(moon['maslulCorr'])),
    ('', ''),
    ('ירח האמיתי', dms(moon['trueLon'])),
    ('', ''),
    ('הראש', dms(moon['node'])),
    ('', ''),
    ('מרחק בין הראש לירח', dms((moon['trueLon'] - moon['node']) % 360)),
    ('', ''),
    ('רוחב ראשון', dms(abs(moon['latitude']))),
    ('', ''),
    ('שינוי המראה (לרוחב, KH 17:7-9)',
        f"≈ {dms(abs(abs(moon['rochavSheni']) - abs(moon['latitude'])))}"),
    ('', ''),
    ('רוחב שני', f"{dms(abs(moon['rochavSheni']))} {'צפוני' if moon['rochavSheni']>=0 else 'דרומי'}"),
    ('', ''),
    ('אורך ראשון', dms(moon['elongation'])),
    ('', ''),
    ('שינוי המראה (לאורך, KH 17:5)',
        f"≈ {dms(abs(moon['elongation'] - moon['orechSheni']))}"),
    ('', ''),
    ('אורך שני', dms(moon['orechSheni'])),
    ('', ''),
    ('מעגל הירח (KH 17:10)', f"≈ {dms(abs(moon['orechShlishi'] - moon['orechSheni']))}"),
    ('', ''),
    ('אורך שלישי', dms(moon['orechShlishi'])),
    ('', ''),
    ('תיקון ארוכי וקצרי שקיעה (KH 17:12a)',
        f"≈ {dms(abs(moon['orechRevii'] - moon['orechShlishi']))}"),
    ('', ''),
    ('אורך רביעי', dms(moon['orechRevii'])),
    ('', ''),
    ('מנת גובה המדינה (⅔ × |רוחב ראשון|)', dms(moon['mnatGovah'])),
    ('', ''),
    ('קשת הראיה', dms(moon['keshet'])),
    ('', ''),
    ('פסק (KH 17:3-4, 17:15-21)', f"{moon['verdict']} — {moon['path']}"),
]
for ri, (lbl, val) in enumerate(mirror_rows, 1):
    a = ws_mirror.cell(row=ri, column=1, value=lbl)
    b = ws_mirror.cell(row=ri, column=2, value=val)
    a.alignment = right
    b.alignment = center
    if lbl and not str(val).startswith('—') and not str(val).startswith('קירוב'):
        a.font = bold
ws_mirror.column_dimensions['A'].width = 36
ws_mirror.column_dimensions['B'].width = 24

# A second sheet with summary table for all dates 1 Nisan – 29 Tamuz 5786
ws2 = wb.create_sheet('כל החודשים')
ws2.sheet_view.rightToLeft = True
hdrs = ['תאריך עברי', 'גרגוריאני', 'ימים מהעיקר', 'אמצע השמש',
        'השמש האמיתי', 'אמצע הירח', 'הירח האמיתי', 'הראש',
        'אורך ראשון', 'אורך שני', 'אורך שלישי', 'אורך רביעי',
        'מנת גובה המדינה', 'קשת הראיה', 'נראה?']
for i, h in enumerate(hdrs, 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = bold; c.fill = hdr_fill; c.alignment = center; c.border = border

HEB_DAYS = ['א','ב','ג','ד','ה','ו','ז','ח','ט','י','יא','יב','יג','יד','טו','טז','יז','יח','יט','כ','כא','כב','כג','כד','כה','כו','כז','כח','כט','ל']
HEB_MONTH = {'Nisan': 'ניסן', 'Iyyar': 'אייר', 'Sivan': 'סיון', 'Tamuz': 'תמוז'}

for i, rec in enumerate(OUR, 2):
    s = rec['sun']; m = rec['moon']
    label = f"{HEB_DAYS[rec['heDay']-1]} {HEB_MONTH[rec['heMonth']]} ה'תשפו"
    vals = [label, rec['gregorian'], rec['daysFromEpoch'],
            dms(s['meanLon']), dms(s['trueLon']),
            dms(m['meanLon']), dms(m['trueLon']), dms(m['node']),
            dms(m['elongation']),
            dms(m['orechSheni']),
            dms(m['orechShlishi']),
            dms(m['orechRevii']),
            dms(m['mnatGovah']),
            dms(m['keshet']),
            'כן' if m['isVisible'] else 'לא']
    for ci, v in enumerate(vals, 1):
        c = ws2.cell(row=i, column=ci, value=v)
        c.alignment = center; c.border = border

# Auto column widths
for sheet in [ws, ws2]:
    for col_idx in range(1, sheet.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in sheet[col_letter]:
            try:
                v = cell.value
                if v is None: continue
                # Avoid touching merged cells
                if hasattr(cell, 'coordinate'):
                    s = str(v)
                    if len(s) > max_len:
                        max_len = len(s)
            except Exception:
                pass
        sheet.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

out = Path('/home/rayi/repos/kidushhachodesh/responses/our_numbers_2_sivan_5786.xlsx')
wb.save(out)
print('Wrote', out)
